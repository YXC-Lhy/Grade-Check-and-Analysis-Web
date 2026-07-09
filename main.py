from __future__ import annotations

import os
import re
import io
import sqlite3
from contextlib import contextmanager
from datetime import date
from typing import Any, Dict, List, Optional

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from openpyxl import load_workbook


DB_PATH = os.path.join(os.path.dirname(__file__), "grades.db")

AUTH_USERNAME = "admin"
AUTH_PASSWORD = "admin123"
FIXED_TOKEN = "fixed-token"

SUBJECTS = [
    ("语文", "yuwen"),
    ("数学", "shuxue"),
    ("英语", "yingyu"),
    ("物理", "wuli"),
    ("化学", "huaxue"),
    ("生物", "shengwu"),
]


def _normalize_header(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\s+", "", s)
    return s


def _is_empty_value(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        s = v.strip()
        return s == "" or s in {"-", "—", "–", "—-", "——", "--", "未上传"}
    return False


def _parse_int(v: Any) -> Optional[int]:
    if _is_empty_value(v):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        raise ValueError(f"不是整数: {v}")
    s = str(v).strip()
    s = s.replace(",", "")
    if s.isdigit():
        return int(s)
    if re.fullmatch(r"[+-]?\d+", s):
        return int(s)
    raise ValueError(f"不是整数: {v}")


def _parse_float(v: Any) -> Optional[float]:
    if _is_empty_value(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if re.fullmatch(r"[+-]?\d+(\.\d+)?", s):
        return float(s)
    raise ValueError(f"不是数字: {v}")


@contextmanager
def db_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        yield conn
    finally:
        conn.close()


def init_db() -> None:
    with db_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                seat_no INTEGER PRIMARY KEY,
                name TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS exams (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                exam_date TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS results (
                exam_id INTEGER NOT NULL,
                seat_no INTEGER NOT NULL,

                yuwen REAL,
                shuxue REAL,
                yingyu REAL,
                wuli REAL,
                huaxue REAL,
                shengwu REAL,
                total REAL,

                yuwen_class_rank INTEGER,
                yuwen_grade_rank INTEGER,
                shuxue_class_rank INTEGER,
                shuxue_grade_rank INTEGER,
                yingyu_class_rank INTEGER,
                yingyu_grade_rank INTEGER,
                wuli_class_rank INTEGER,
                wuli_grade_rank INTEGER,
                huaxue_class_rank INTEGER,
                huaxue_grade_rank INTEGER,
                shengwu_class_rank INTEGER,
                shengwu_grade_rank INTEGER,

                total_class_rank INTEGER,
                total_grade_rank INTEGER,

                PRIMARY KEY (exam_id, seat_no),
                FOREIGN KEY (exam_id) REFERENCES exams(id) ON DELETE CASCADE,
                FOREIGN KEY (seat_no) REFERENCES students(seat_no) ON DELETE CASCADE
            );
            """
        )


class LoginIn(BaseModel):
    username: str
    password: str


class LoginOut(BaseModel):
    token: str
    token_type: str = "bearer"


def require_token(authorization: Optional[str] = Header(default=None)) -> None:
    if not authorization:
        raise HTTPException(status_code=401, detail="缺少 Authorization")
    m = re.fullmatch(r"Bearer\s+(.+)", authorization.strip(), flags=re.IGNORECASE)
    if not m:
        raise HTTPException(status_code=401, detail="Authorization 格式必须为 Bearer <token>")
    token = m.group(1).strip()
    if token != FIXED_TOKEN:
        raise HTTPException(status_code=401, detail="token 无效")


app = FastAPI(title="成绩分析站后端", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _frontend_file(name: str) -> str:
    return os.path.join(os.path.dirname(__file__), name)


@app.get("/", include_in_schema=False)
def _root() -> FileResponse:
    return FileResponse(_frontend_file("login.html"))


@app.get("/login.html", include_in_schema=False)
def _login_page() -> FileResponse:
    return FileResponse(_frontend_file("login.html"))


@app.get("/query.html", include_in_schema=False)
def _query_page() -> FileResponse:
    return FileResponse(_frontend_file("query.html"))


@app.get("/upload.html", include_in_schema=False)
def _upload_page() -> FileResponse:
    return FileResponse(_frontend_file("upload.html"))


@app.on_event("startup")
def _startup() -> None:
    init_db()


@app.post("/login", response_model=LoginOut)
def login(body: LoginIn) -> LoginOut:
    if body.username != AUTH_USERNAME or body.password != AUTH_PASSWORD:
        raise HTTPException(status_code=401, detail="账号或密码错误")
    return LoginOut(token=FIXED_TOKEN)


def _expected_headers() -> List[str]:
    headers: List[str] = ["座号", "姓名"]
    headers.extend([s[0] for s in SUBJECTS])
    headers.append("总分")
    for cn, _ in SUBJECTS:
        headers.append(f"{cn}班排")
        headers.append(f"{cn}段排")
    headers.append("总分班排")
    headers.append("总分段排")
    return headers


def _get_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _normalize_header(ws.cell(row=1, column=col).value)
        if not key:
            continue
        header_map[key] = col
    return header_map


@app.post("/exams/import")
async def import_exam(
    name: str = Form(...),
    exam_date: str = Form(...),
    file: UploadFile = File(...),
    _: None = Depends(require_token),
) -> Dict[str, Any]:
    try:
        d = date.fromisoformat(exam_date)
    except Exception:
        raise HTTPException(status_code=400, detail="exam_date 必须为 YYYY-MM-DD")

    content = await file.read()
    try:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Excel 文件无法解析（仅支持 xlsx）")

    ws = wb.worksheets[0]
    header_map = _get_header_map(ws)

    expected = [_normalize_header(h) for h in _expected_headers()]
    missing = [h for h in expected if h not in header_map]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={"message": "表头不完整，未找到全部 23 列", "missing_headers": missing},
        )

    imported = 0
    with db_conn() as conn:
        try:
            conn.execute("BEGIN")
            exists = conn.execute("SELECT 1 FROM exams WHERE name = ?", (name,)).fetchone()
            if exists:
                raise HTTPException(status_code=400, detail="考试名称已存在")
            cur = conn.execute(
                "INSERT INTO exams (name, exam_date) VALUES (?, ?)",
                (name, d.isoformat()),
            )
            exam_id = int(cur.lastrowid)

            for r in range(2, ws.max_row + 1):
                raw_seat = ws.cell(row=r, column=header_map["座号"]).value
                raw_name = ws.cell(row=r, column=header_map["姓名"]).value

                if _is_empty_value(raw_seat) and _is_empty_value(raw_name):
                    continue

                try:
                    seat_no = _parse_int(raw_seat)
                except Exception:
                    raise HTTPException(status_code=400, detail=f"第 {r} 行：座号不是整数")
                if seat_no is None:
                    raise HTTPException(status_code=400, detail=f"第 {r} 行：座号为空")

                if _is_empty_value(raw_name):
                    raise HTTPException(status_code=400, detail=f"第 {r} 行：姓名为空")
                student_name = str(raw_name).strip()

                conn.execute(
                    """
                    INSERT INTO students (seat_no, name) VALUES (?, ?)
                    ON CONFLICT(seat_no) DO UPDATE SET name = excluded.name
                    """,
                    (seat_no, student_name),
                )

                scores: Dict[str, Optional[float]] = {}
                ranks: Dict[str, Optional[int]] = {}
                for cn, key in SUBJECTS:
                    scores[key] = _parse_float(ws.cell(row=r, column=header_map[cn]).value)
                    ranks[f"{key}_class_rank"] = _parse_int(
                        ws.cell(row=r, column=header_map[f"{cn}班排"]).value
                    )
                    ranks[f"{key}_grade_rank"] = _parse_int(
                        ws.cell(row=r, column=header_map[f"{cn}段排"]).value
                    )

                total_score = _parse_float(ws.cell(row=r, column=header_map["总分"]).value)
                total_class_rank = _parse_int(ws.cell(row=r, column=header_map["总分班排"]).value)
                total_grade_rank = _parse_int(ws.cell(row=r, column=header_map["总分段排"]).value)

                conn.execute(
                    """
                    INSERT INTO results (
                        exam_id, seat_no,
                        yuwen, shuxue, yingyu, wuli, huaxue, shengwu, total,
                        yuwen_class_rank, yuwen_grade_rank,
                        shuxue_class_rank, shuxue_grade_rank,
                        yingyu_class_rank, yingyu_grade_rank,
                        wuli_class_rank, wuli_grade_rank,
                        huaxue_class_rank, huaxue_grade_rank,
                        shengwu_class_rank, shengwu_grade_rank,
                        total_class_rank, total_grade_rank
                    ) VALUES (
                        ?, ?,
                        ?, ?, ?, ?, ?, ?, ?,
                        ?, ?,
                        ?, ?,
                        ?, ?,
                        ?, ?,
                        ?, ?,
                        ?, ?,
                        ?, ?
                    )
                    """,
                    (
                        exam_id,
                        seat_no,
                        scores["yuwen"],
                        scores["shuxue"],
                        scores["yingyu"],
                        scores["wuli"],
                        scores["huaxue"],
                        scores["shengwu"],
                        total_score,
                        ranks["yuwen_class_rank"],
                        ranks["yuwen_grade_rank"],
                        ranks["shuxue_class_rank"],
                        ranks["shuxue_grade_rank"],
                        ranks["yingyu_class_rank"],
                        ranks["yingyu_grade_rank"],
                        ranks["wuli_class_rank"],
                        ranks["wuli_grade_rank"],
                        ranks["huaxue_class_rank"],
                        ranks["huaxue_grade_rank"],
                        ranks["shengwu_class_rank"],
                        ranks["shengwu_grade_rank"],
                        total_class_rank,
                        total_grade_rank,
                    ),
                )
                imported += 1

            conn.commit()
        except HTTPException:
            conn.rollback()
            raise
        except ValueError as e:
            conn.rollback()
            raise HTTPException(status_code=400, detail=str(e))
        except Exception:
            conn.rollback()
            raise HTTPException(status_code=500, detail="导入失败")

    return {"exam_id": exam_id, "name": name, "exam_date": d.isoformat(), "imported": imported}


@app.get("/meta")
def meta() -> Dict[str, Any]:
    with db_conn() as conn:
        students = conn.execute("SELECT seat_no, name FROM students ORDER BY seat_no").fetchall()
        exams = conn.execute("SELECT id, name, exam_date FROM exams ORDER BY exam_date, id").fetchall()
    return {
        "students": [{"seat_no": int(r["seat_no"]), "name": r["name"]} for r in students],
        "exams": [{"id": int(r["id"]), "name": r["name"], "exam_date": r["exam_date"]} for r in exams],
    }


@app.get("/students/{seat_no}/results")
def student_results(seat_no: int, _: None = Depends(require_token)) -> Dict[str, Any]:
    with db_conn() as conn:
        student = conn.execute("SELECT seat_no, name FROM students WHERE seat_no = ?", (seat_no,)).fetchone()
        if not student:
            raise HTTPException(status_code=404, detail="学生不存在")
        rows = conn.execute(
            """
            SELECT
                e.id AS exam_id, e.name AS exam_name, e.exam_date,
                r.*
            FROM results r
            JOIN exams e ON e.id = r.exam_id
            WHERE r.seat_no = ?
            ORDER BY e.exam_date, e.id
            """,
            (seat_no,),
        ).fetchall()

    exams: List[Dict[str, Any]] = []
    for row in rows:
        subjects: Dict[str, Any] = {}
        for cn, key in SUBJECTS:
            subjects[cn] = {
                "score": row[key],
                "class_rank": row[f"{key}_class_rank"],
                "grade_rank": row[f"{key}_grade_rank"],
            }
        exams.append(
            {
                "exam_id": int(row["exam_id"]),
                "exam_name": row["exam_name"],
                "exam_date": row["exam_date"],
                "subjects": subjects,
                "total": {
                    "score": row["total"],
                    "class_rank": row["total_class_rank"],
                    "grade_rank": row["total_grade_rank"],
                },
            }
        )

    return {"student": {"seat_no": int(student["seat_no"]), "name": student["name"]}, "exams": exams}


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    os.system("")
    uvicorn.run(app, host="127.0.0.1", port=12345)

